'''
IMPORT PYTHON FILE: TESTING IN PROGRESS
'''

from string import ascii_lowercase
from biological_samples_database.model.sample import Antigen, CellLine, Mosquito, Other, Pbmc, Peptide, Plasma, Rna, Serum, Supernatant, Vial, VirusCulture, VirusIsolation
from biological_samples_database.model.storage import Box, BoxType, Building, Freezer, FreezerType, Room, Shelf 
from biological_samples_database.database import create_new_session, engine
import openpyxl
import os
from datetime import datetime
from sqlalchemy.orm import sessionmaker

sess = sessionmaker()
sess.configure(bind=engine)
 



def import_excel(path):
    dataframe = openpyxl.load_workbook(path) # need to change to import button
    dataframe1 = dataframe.active
    fill_box(dataframe1)
    os.remove(path) # remove the file so we don't store them

def fill_box(all_data):
    box_table = [] # Holds the data to most of the excel sheet
    for row in range(0, 6):
        for col in all_data.iter_cols(2, 2):
            box_table.append(col[row].value)
        if row == 3:
            for col in all_data.iter_cols(1, 1):
                fridge_type = (col[row].value)
    start_new = sess()
    obj2 = start_new.query(Box).all()
    count1 = start_new.query(Box).count()
    count2 = 0
    for boxs in obj2:
        count2 += 1
        if (boxs.label == box_table[0] or boxs.label == str("Box Position: " + str(box_table[4]) + " | Name: " + str(box_table[0]))):
            print("ERROR: BOX NAME ALREADY IN DATABASE") 
            print("CHECK THAT YOU HAVEN'T ALREADY IMPORTED THIS FILE")
            break
        elif (count1 == count2):
            box(box_table, fridge_type)
            box_sess = sess()
            if fridge_type == "Tower ID:":
                obj = box_sess.query(Box).filter(Box.label == str("Box Position: " + str(box_table[4]) + " | Name: " + str(box_table[0])))
            else: 
                obj = box_sess.query(Box).filter(Box.label == box_table[0])
            for k in obj:
                box_id = k.id

            add_vials(all_data, box_id) 


def get_fzrType(fridge_type):
    fridge_sess = sess()
    if fridge_type == "Tower ID:":
        q = fridge_sess.query(FreezerType).filter(FreezerType.name == 'LN2')
        for val in q:
            return val.id
    else:
        q = fridge_sess.query(FreezerType).filter(FreezerType.name == '-80c')
        for val in q:
            return val.id



def alnum_to_coord(alnum, boxid):
    ALPHA_MAP = {char:index for index, char in enumerate(ascii_lowercase, start = 1)}
    alpha = ALPHA_MAP[alnum[0].lower()]
    num = int(alnum[1:])-1
    with create_new_session() as session:
        box = session.query(
            Box
        ).filter(
            Box.id == boxid
        )
        for val in box:
            box_type = val.box_type
        box_type = session.query(
            BoxType
        ).filter(
            BoxType.id == box_type
        )
        if box_type:
            for val in box_type:
                width = val.width
            return str(alpha+width*num)
    raise ValueError("Box id does not point to a Box")

def datetime_conversion(date):
    if date == None:
        return None
    elif type(date) == datetime:
        return date
    x = date.split('/')
    if(int(x[0]) > 12 or int(x[1]) > 31): # Date inputted wrong
        return None
    date_time = datetime(year=int(x[2]), month = int(x[0]), day = int(x[1]))
    return date_time


# Get Freezer ID or create new freezer if freezer is not in database currently
def new_freezer(box_table, fridge_type):
    box_sess = sess()
    q = box_sess.query(Freezer).all()
    p = box_sess.query(Freezer).count()
    counter = 0

    
    if fridge_type == "Tower ID:":
        ln2_query = box_sess.query(Freezer).filter(Freezer.name == 'LN2 Freezer').one_or_none()
        if ln2_query is None:
            #Create new LN2 freezer and room
            NewRoom = Room()
            build_query = box_sess.query(Building).first()
            NewRoom.building_id = build_query.id
            NewRoom.name = 'LN2 Room'
            newsess2 = sess()
            newsess2.add(NewRoom)
            newsess2.commit()
            obj2 = box_sess.query(Room).filter(Room.name == 'LN2 Room')
            for val2 in obj2:
                ln2_roomID = val2.id


            NewFreezer = Freezer()
            NewFreezer.name = 'LN2 Freezer'
            NewFreezer.freezer_type = get_fzrType(fridge_type) #sets to the freezer type before
            NewFreezer.room_id = ln2_roomID #sets to the same room as before
            newsess = sess()
            newsess.add(NewFreezer)
            newsess.commit()
            obj = box_sess.query(Freezer).filter(Freezer.name == 'LN2 Freezer')
            for j in obj:
                return j.id
        else:
            return ln2_query.id


    for i in q:
        counter +=1
        if(i.name == box_table[3]): #if statement off name not ID atm
            return i.id
        elif (p <= counter):
            NewFreezer = Freezer()
            NewFreezer.name = box_table[3]
            NewFreezer.freezer_type = get_fzrType(fridge_type) #sets to the freezer type before
            NewFreezer.room_id = i.room_id #sets to the same room as before
            newsess = sess()
            newsess.add(NewFreezer)
            newsess.commit()
            obj = box_sess.query(Freezer).filter(Freezer.name == box_table[3])
            for j in obj:
                return j.id

# Get the type of the box from database
def box_type(box_table):
    new_sess = sess()
    obj = new_sess.query(BoxType).all()
    for type1 in obj:
        
        if (type1.name == "9x9"):
            nineBynine = type1.id
        elif(type1.name == "10x10"):
            tenByten = type1.id
        elif(type1.name == "Wax Box (Standard)"):
            WBS = type1.id
        elif(type1.name == "Wax Box (5ml)"):
            WB5ml = type1.id
        elif(type1.name == "Wax Box (Large)"):
            WBL = type1.id
    if(box_table[1] == "Wax Box standard"):
        return WBS
    elif(box_table[1] == "Wax Box (5ml)"): #ToDo fix these up when needed
        return WB5ml
    elif(box_table[1] == "Wax Box large"):
        return WBL
    elif(box_table[1] == "10x10"): 
        return tenByten
    elif(box_table[1] == "9x9"): 
        return nineBynine
    

#Get the shelf that the box is on or make new shelf if not in database
def shelf_box(freezer_id, box_table, fridge_type):
    shelf_sess = sess()
    obj = shelf_sess.query(Shelf).filter(Shelf.freezer_id == freezer_id)
    count1 = shelf_sess.query(Shelf).filter(Shelf.freezer_id == freezer_id).count()
    counter = 0


    if count1 == 0:
        shelf = Shelf()
        shelf.freezer_id = freezer_id
        if fridge_type == "Tower ID:":
            shelf.name = box_table[3]
        else:
            shelf.name = box_table[4]
        new_session = sess()
        new_session.add(shelf)
        new_session.commit()
        if fridge_type == "Tower ID:":
            obj1 = shelf_sess.query(Shelf).filter(Shelf.name == box_table[3])
        else:
            obj1 = shelf_sess.query(Shelf).filter(Shelf.name == box_table[4])
        for j in obj1:
            return j.id
    
    else:
        for shelf1 in obj:
            counter +=1
            if(str(shelf1.name) == str(box_table[4]) or str(shelf1.name) == str(box_table[3])):
                return shelf1.id

            elif(count1 == counter):
                shelf = Shelf()
                shelf.freezer_id = freezer_id
                if fridge_type == "Tower ID:":
                    shelf.name = box_table[3]
                else:
                    shelf.name = box_table[4]
                new_session = sess()
                new_session.add(shelf)
                new_session.commit()
                if fridge_type == "Tower ID:":
                    obj1 = shelf_sess.query(Shelf).filter(Shelf.name == box_table[3])
                else:
                    obj1 = shelf_sess.query(Shelf).filter(Shelf.name == box_table[4])
                for j in obj1:
                    return j.id

# Fill the box table
def box(box_table, fridge_type):
    box_sess = sess()
    box = Box()
    if fridge_type == "Tower ID:":
        box.label = str("Box Position: " + str(box_table[4]) + " | Name: " + str(box_table[0]))
    else:
        box.label = box_table[0]
    box.owner = box_table[5]
    box.box_type = box_type(box_table)
    freezer_id = new_freezer(box_table, fridge_type)
    box.freezer_id = freezer_id
    box.shelf_id = shelf_box(freezer_id, box_table, fridge_type)
    box_sess.add(box)
    box_sess.commit()

def antigen(data_row, box_id):
    Sess = sess()
    new_entry = Antigen()
    new_entry.pathwest_id = data_row[2]
    new_entry.lab_id = data_row[3]
    new_entry.batch_number = data_row[7]
    new_entry.passage_number = data_row[8]
    new_entry.lot_number = data_row[12]
    if isinstance(data_row[0],str) and str.isalnum(data_row[0]):
        new_entry.position = alnum_to_coord(data_row[0], box_id)
    else:
        new_entry.position = data_row[0]
    new_entry.sample_date = datetime_conversion(data_row[5])
    #new_entry.volume_ml = data_row[13]
    new_entry.box_id = box_id
    new_entry.user_id = data_row[15]
    new_entry.notes = data_row[16]
    Sess.add(new_entry)
    Sess.commit()

def cell_line(data_row, box_id):
    Sess = sess()
    new_entry = CellLine()
    new_entry.lab_id = data_row[3]
    new_entry.cell_type = data_row[4]
    new_entry.passage_number = data_row[8]
    #new_entry.cell_count = data_row[9] data type needs to be changed
    new_entry.growth_media = data_row[10]
    new_entry.vial_source = data_row[11]
    new_entry.lot_number = data_row[12]
    if isinstance(data_row[0],str) and str.isalnum(data_row[0]):
        new_entry.position = alnum_to_coord(data_row[0], box_id)
    else:
        new_entry.position = data_row[0]
    new_entry.sample_date = datetime_conversion(data_row[5])
    #new_entry.volume_ml = data_row[13]
    new_entry.box_id = box_id
    new_entry.user_id = data_row[15]
    new_entry.notes = data_row[16]
    Sess.add(new_entry)
    Sess.commit()

def mosquito(data_row, box_id):
    Sess = sess()
    new_entry = Mosquito()
    new_entry.lab_id = data_row[3]
    if isinstance(data_row[0],str) and str.isalnum(data_row[0]):
        new_entry.position = alnum_to_coord(data_row[0], box_id)
    else:
        new_entry.position = data_row[0]
    new_entry.sample_date = datetime_conversion(data_row[5])
    #new_entry.volume_ml = data_row[13]
    new_entry.box_id = box_id
    new_entry.user_id = data_row[15]
    new_entry.notes = data_row[16]
    Sess.add(new_entry)
    Sess.commit()

def pbmc(data_row, box_id):
    Sess = sess()
    new_entry = Pbmc()
    new_entry.lab_id = data_row[3]
    new_entry.visit_number = data_row[6]
    #new_entry.cell_count = data_row[9] data type needs to be changed
    new_entry.patient_code = data_row[14]
    if isinstance(data_row[0],str) and str.isalnum(data_row[0]):
        new_entry.position = alnum_to_coord(data_row[0], box_id)
    else:
        new_entry.position = data_row[0]
    new_entry.sample_date = datetime_conversion(data_row[5])
    #new_entry.volume_ml = data_row[13]
    new_entry.box_id = box_id
    new_entry.user_id = data_row[15]
    new_entry.notes = data_row[16]
    Sess.add(new_entry)
    Sess.commit()

def peptide(data_row, box_id):
    Sess = sess()
    new_entry = Peptide()
    new_entry.lab_id = data_row[3]
    new_entry.cell_type = data_row[4]
    new_entry.batch_number = data_row[7]
    new_entry.vial_source = data_row[11]
    new_entry.lot_number = data_row[12]
    if isinstance(data_row[0],str) and str.isalnum(data_row[0]):
        new_entry.position = alnum_to_coord(data_row[0], box_id)
    else:
        new_entry.position = data_row[0]
    new_entry.sample_date = datetime_conversion(data_row[5])
    #new_entry.volume_ml = data_row[13]
    new_entry.box_id = box_id
    new_entry.user_id = data_row[15]
    new_entry.notes = data_row[16]
    Sess.add(new_entry)
    Sess.commit()

def plasma(data_row, box_id):
    Sess = sess()
    new_entry = Plasma()
    new_entry.lab_id = data_row[3]
    new_entry.visit_number = data_row[6]
    if isinstance(data_row[0],str) and str.isalnum(data_row[0]):
        new_entry.position = alnum_to_coord(data_row[0], box_id)
    else:
        new_entry.position = data_row[0]
    new_entry.sample_date = datetime_conversion(data_row[5])
    #new_entry.volume_ml = data_row[13]
    new_entry.box_id = box_id
    new_entry.user_id = data_row[15]
    new_entry.notes = data_row[16]
    Sess.add(new_entry)
    Sess.commit()

def rna(data_row, box_id):
    Sess = sess()
    new_entry = Rna()
    new_entry.pathwest_id = data_row[2]
    new_entry.lab_id = data_row[3]
    new_entry.batch_number = data_row[7]
    new_entry.lot_number = data_row[12]
    if isinstance(data_row[0],str) and str.isalnum(data_row[0]):
        new_entry.position = alnum_to_coord(data_row[0], box_id)
    else:
        new_entry.position = data_row[0]
    new_entry.sample_date = datetime_conversion(data_row[5])
    #new_entry.volume_ml = data_row[13]
    new_entry.box_id = box_id
    new_entry.user_id = data_row[15]
    new_entry.notes = data_row[16]
    Sess.add(new_entry)
    Sess.commit()

def serum(data_row, box_id):
    Sess = sess()
    new_entry = Serum()
    new_entry.pathwest_id = data_row[2]
    new_entry.lab_id = data_row[3]
    if isinstance(data_row[0],str) and str.isalnum(data_row[0]):
        new_entry.position = alnum_to_coord(data_row[0], box_id)
    else:
        new_entry.position = data_row[0]
    new_entry.sample_date = datetime_conversion(data_row[5])
    #new_entry.volume_ml = data_row[13]
    new_entry.box_id = box_id
    new_entry.user_id = data_row[15]
    new_entry.notes = data_row[16]
    Sess.add(new_entry)
    Sess.commit()

def supernatant(data_row, box_id):
    Sess = sess()
    new_entry = Supernatant()
    new_entry.lab_id = data_row[3]
    if isinstance(data_row[0],str) and str.isalnum(data_row[0]):
        new_entry.position = alnum_to_coord(data_row[0], box_id)
    else:
        new_entry.position = data_row[0]
    new_entry.sample_date = datetime_conversion(data_row[5])
    #new_entry.volume_ml = data_row[13]
    new_entry.box_id = box_id
    new_entry.user_id = data_row[15]
    new_entry.notes = data_row[16]
    Sess.add(new_entry)
    Sess.commit()

def virus_culture(data_row, box_id):
    Sess = sess()
    new_entry = VirusCulture()
    new_entry.pathwest_id = data_row[2]
    new_entry.lab_id = data_row[3]
    new_entry.batch_number = data_row[7]
    new_entry.passage_number = data_row[8]
    new_entry.growth_media = data_row[10]
    if isinstance(data_row[0],str) and str.isalnum(data_row[0]):
        new_entry.position = alnum_to_coord(data_row[0], box_id)
    else:
        new_entry.position = data_row[0]
    new_entry.sample_date = datetime_conversion(data_row[5])
    #new_entry.volume_ml = data_row[13]
    new_entry.box_id = box_id
    new_entry.user_id = data_row[15]
    new_entry.notes = data_row[16]
    Sess.add(new_entry)
    Sess.commit()

def virus_isolation(data_row, box_id):
    Sess = sess()
    new_entry = VirusIsolation()
    new_entry.box_id = box_id
    if isinstance(data_row[0],str) and str.isalnum(data_row[0]):
        new_entry.position = alnum_to_coord(data_row[0], box_id)
    else:
        new_entry.position = data_row[0]
    new_entry.pathwest_id = data_row[2]
    new_entry.lab_id = data_row[3]
    new_entry.sample_date = datetime_conversion(data_row[5])
    new_entry.batch_number = data_row[7]
    new_entry.passage_number = data_row[8]
    new_entry.growth_media = data_row[10]
    #new_entry.volume_ml = data_row[13]
    new_entry.user_id = data_row[15]
    new_entry.notes = data_row[16]
    Sess.add(new_entry)
    Sess.commit()

def other(data_row, box_id):
    Sess = sess()
    new_entry = Other()
    new_entry.lab_id = data_row[3]
    if isinstance(data_row[0],str) and str.isalnum(data_row[0]):
        new_entry.position = alnum_to_coord(data_row[0], box_id)
    else:
        new_entry.position = data_row[0]
    new_entry.sample_date = datetime_conversion(data_row[5])
    #new_entry.volume_ml = data_row[13] Leaving this out because field type isn't a match... same as cell_count
    new_entry.box_id = box_id
    new_entry.user_id = data_row[15]
    new_entry.notes = data_row[16]
    Sess.add(new_entry)
    Sess.commit()

def add_vials(all_data, box_id):    
    
    for row in range(7, all_data.max_row):
        next = []
        for col in all_data.iter_cols(1, all_data.max_column):
            next.append(col[row].value)
        if next[1] == "Virus Isolation":
            virus_isolation(next, box_id)
        elif next[1] == "Cell Line":
            cell_line(next, box_id)
        elif next[1] == "Mosquito":
            mosquito(next, box_id)
        elif next[1] == "PBMC":
            pbmc(next, box_id)
        elif next[1] == "Plasma":
            plasma(next, box_id)
        elif next[1] == "Serum":
            serum(next, box_id)
        elif next[1] == "Virus Culture":
            virus_culture(next, box_id)
        elif next[1] == "Supernatant":
            supernatant(next, box_id)
        elif next[1] == "RNA":
            rna(next, box_id)
        elif next[1] == "Antigen":
            antigen(next, box_id)
        elif next[1] == "Peptide":
            peptide(next, box_id)
        elif next[1] == "Other":
            other(next, box_id)
        else: #IF IT DOESN'T MATCH ANY OF THE SAMPLE TYPES ABOVE IT LEAVES THE POSITIONS BLANK
            continue    

